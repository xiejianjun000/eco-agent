from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws1 = wb.active
ws1.title = '评分汇总'

header_font = Font(name='SimHei', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1A5276')
cell_font = Font(name='SimSun', size=10)
title_font = Font(name='SimHei', bold=True, size=14, color='1A5276')
thin_border = Border(
    left=Side(style='thin', color='999999'),
    right=Side(style='thin', color='999999'),
    top=Side(style='thin', color='999999'),
    bottom=Side(style='thin', color='999999'),
)

ws1.merge_cells('A1:G1')
ws1['A1'] = '生态环境行政执法案卷评查评分汇总表'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')
ws1['A2'] = '案卷编号: 娄环罚(冷)[2026]2号 | 评查日期: 2026-08-04'
ws1['A2'].font = Font(name='SimSun', size=10, color='666666')

headers = ['评查维度', '满分', '得分', '扣分', '扣分项数', '得分率', '等级']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border

data = [
    ['一、立案审批', 10, 10, 0, 0, '100%', '优秀'],
    ['二、调查取证', 25, 23, 2, 1, '92%', '合格'],
    ['三、告知与陈述申辩', 15, 15, 0, 0, '100%', '优秀'],
    ['四、法制审核与集体讨论', 10, 7, 3, 1, '70%', '需改进'],
    ['五、行政处罚决定', 15, 14, 1, 1, '93%', '合格'],
    ['六、文书送达与归档', 10, 9, 1, 1, '90%', '合格'],
    ['七、法律适用', 15, 14, 1, 1, '93%', '合格'],
]

for r, row in enumerate(data, 5):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.font = cell_font; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if c == 7:
            if val == '优秀': cell.font = Font(name='SimSun', size=10, color='008000', bold=True)
            elif val == '需改进': cell.font = Font(name='SimSun', size=10, color='CC0000', bold=True)

total_row = 12
total_data = ['合 计', 100, 92, 8, 5, '92%', '合格']
for c, val in enumerate(total_data, 1):
    cell = ws1.cell(row=total_row, column=c, value=val)
    cell.font = Font(name='SimHei', bold=True, size=11); cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill('solid', fgColor='E8F0FE')

ws1.merge_cells('A14:G14')
ws1['A14'] = '25项一票否决扫描结果: 未命中 (1项V09法制审核需核实，但不足以触发否决)'
ws1['A14'].font = Font(name='SimHei', size=11, color='008000')

widths1 = [28, 8, 8, 8, 10, 8, 10]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# === Sheet 2 ===
ws2 = wb.create_sheet('扣分明细')
ws2.merge_cells('A1:H1')
ws2['A1'] = '扣分明细表'
ws2['A1'].font = title_font; ws2['A1'].alignment = Alignment(horizontal='center')

headers2 = ['序号', '维度', '编号', '问题描述', '关联法条/规范', '扣分', '等级', '整改建议']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border

deductions = [
    [1, '调查取证', 'D1', '询问笔录中刘忠于对部分问题回答不清楚，调查深度可进一步加强', '《环境行政处罚办法》第26条', 2, '低', '对重要事实尽可能向多名知情人核实'],
    [2, '法制审核与集体讨论', 'D2', '案卷中未见独立法制审核意见书，法制审核程序材料不完整', '《行政处罚法》第58条', 3, '中', '补充法制审核意见书或书面说明审核情况'],
    [3, '行政处罚决定', 'D3', '决定书罚款金额文字表述OCR模糊(肆万玖仟元)', '案卷文书规范', 1, '低', '确保原件清晰，必要时重新制作'],
    [4, '文书送达与归档', 'D4', '送达回证签收日期OCR模糊不可辨认', '案卷归档管理规定', 1, '低', '确保送达回证原件清晰可读'],
    [5, '法律适用', 'D5', '未按法典过渡期要求标注法典对应条文(双标注)', '生态环境法典过渡期管理指南', 1, '低', '增加法典对应条款标注'],
]

severity_fills = {'高': PatternFill('solid', fgColor='FFD7D7'), '中': PatternFill('solid', fgColor='FFF3CD'), '低': PatternFill('solid', fgColor='D5F5E3')}

for r, row in enumerate(deductions, 4):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = cell_font; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True) if c != 4 else Alignment(wrap_text=True)
        if c == 7 and val in severity_fills:
            cell.fill = severity_fills[val]

widths2 = [6, 20, 10, 42, 24, 6, 8, 24]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# === Sheet 3 ===
ws3 = wb.create_sheet('否决扫描详情')
ws3.merge_cells('A1:F1')
ws3['A1'] = '25项一票否决扫描详情'
ws3['A1'].font = title_font; ws3['A1'].alignment = Alignment(horizontal='center')

headers3 = ['编号', '类别', '名称', '关键词匹配', '置信度', '结果']
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border

veto_data = [
    ['V01-V08', '程序类', '告知/单人/亮证/时效等', '全部合规', '高', '未触发'],
    ['V09', '程序类', '法制审核缺失', '案审会提及但未见独立审核书', '中', '需核实'],
    ['V10', '程序类', '听证期限不足', '罚款2.9万不达听证标准', '高', '未触发'],
    ['V11-V15', '证据类', '孤证/采样/监测/签字等', '6类10项证据链完整', '高', '未触发'],
    ['V16-V18', '法律适用类', '废止法/条文错误/裁量不当', '过渡期现行有效/计算正确', '高', '未触发'],
    ['V19-V20', '定性移送类', '应移未移/查封扣押', '不涉及移送公安/不涉及查封', '高', '未触发'],
    ['V21-V22', '主体管辖类', '主体不适格/管辖问题', '适格主体/有管辖权', '高', '未触发'],
    ['V23-V25', '文书期限类', '超期/要素缺失/强执违法', '要素齐全/不涉及强执', '高', '未触发'],
]

for r, row in enumerate(veto_data, 4):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.font = cell_font; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

widths3 = [12, 14, 20, 32, 8, 12]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# === Sheet 4: 法律依据链 ===
ws4 = wb.create_sheet('法律依据链')
ws4.merge_cells('A1:E1')
ws4['A1'] = '法律依据链 (按效力层级排列)'
ws4['A1'].font = title_font; ws4['A1'].alignment = Alignment(horizontal='center')

headers4 = ['效力层级', '法规名称', '条款', '条文内容(摘要)', '作用']
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=3, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border

legal = [
    ['法律', '《大气污染防治法》', '第48条第2款', '工业企业应采取密闭、围挡、遮盖等措施减少粉尘排放', '禁止性规范(义务条款)'],
    ['法律', '《大气污染防治法》', '第108条第5项', '矿产开采企业未采取防尘措施的，处2万-20万罚款', '罚则'],
    ['法律', '《行政处罚法》', '第28条', '实施行政处罚时应责令改正或限期改正', '责令改正依据'],
    ['法律', '《行政处罚法》', '第44条', '处罚前应告知事实、理由、依据及陈述申辩权利', '告知程序依据'],
    ['法律', '《行政处罚法》', '第58条', '特定情形须经法制审核', '法制审核依据'],
    ['法律', '《行政处罚法》', '第72条第1款第1项', '逾期不缴罚款可每日加处3%', '逾期加罚依据'],
    ['法律', '《行政复议法》', '第9条', '60日内申请行政复议', '复议权利告知'],
    ['法律', '《行政诉讼法》', '第46条', '6个月内提起行政诉讼', '诉讼权利告知'],
    ['地方规范', '《湖南省裁量权基准(2021版)》', '表13通用裁量表', 'Y=10%; 罚款=[Y+Px(1-Y)]x20万=2.9万', '裁量计算依据'],
    ['法律(过渡)', '《生态环境法典》', '(2026.8.15施行)', '第1242条废止10部单行法; 对应条款待标注', '法典过渡期标注'],
]

for r, row in enumerate(legal, 4):
    for c, val in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.font = cell_font; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True) if c != 4 else Alignment(wrap_text=True)

widths4 = [12, 28, 16, 48, 24]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

out = 'C:/Users/Administrator/WorkBuddy/执法督察评查专家团/评查明细_娄环罚(冷)〔2026〕2号_20260804.xlsx'
wb.save(out)
print(f'Saved: {out}')
